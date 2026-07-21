#!/usr/bin/env python3
"""
bootstrap_from_workbook.py — ONE-SHOT seed of the web tracker DB from the
laptop's Excel workbook. Phase 2 of docs/PLAN_web_tracker.md.

    python3 scripts/bootstrap_from_workbook.py --dry-run     # report, no writes
    python3 scripts/bootstrap_from_workbook.py               # pull + import
    python3 scripts/bootstrap_from_workbook.py --workbook x.xlsx   # use a local file

This runs exactly once, at the moment the Pi takes over as system of record.
After it, the laptop's workbook is a stale artifact.

WHAT GETS IMPORTED, AND WHY IT IS THREE SOURCES

  matched_jobs.xlsx  -> live rows, hand-typed columns preserved.
                        The workbook is the truth for what you still care
                        about: it holds your edits AND reflects every row you
                        ever deleted.

  matched_jobs.csv   -> rows NOT in the workbook become TOMBSTONES.
                        The CSV is cumulative and still contains everything you
                        hand-deleted. Import it as live data and the tracker
                        instantly refills with months of rejected listings;
                        skip it entirely and the pipeline re-adds them all on
                        its next cycle, because nothing records that they were
                        deleted. Tombstoning is what makes the deletion stick.

  pruned_keys.txt    -> TOMBSTONES, same reasoning.
                        This file exists today for exactly this purpose: it is
                        the suppress list export_workbook consults so pruned
                        rows are never re-added.

REFUSES TO RUN TWICE. A second run after you have edited anything in the web
UI would overwrite live work with a stale laptop snapshot, so a non-empty jobs
table aborts the run — there is deliberately no --force. Recovering from a bad
import means deleting data/tracker.db and starting over, which is safe because
this is the only writer that has run at that point.
"""

import argparse
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

# export_workbook exits with a pip hint if openpyxl is missing, so import it
# first. It owns the workbook schema — COLUMNS and the HYPERLINK parser are
# taken from there rather than restated, so a column rename can't silently
# desync this importer.
from export_workbook import COLUMNS, HEADERS, XLSX_PATH, strip_hyperlink
from openpyxl import load_workbook

import db
from matches import TS_FORMAT, read_matches, row_key
from paths import DATA_DIR
from remote import load_local_config, remote_base, scp

CSV_PATH         = DATA_DIR / "matched_jobs.csv"
PRUNED_KEYS_PATH = DATA_DIR / "pruned_keys.txt"

# export_workbook auto-fills Application ID with "." as an overflow spacer, so
# "." is not a user value. Same set prune_workbook uses to decide emptiness.
PLACEHOLDERS = {"", "."}

# Workbook header -> jobs column, for the columns the user owns.
USER_HEADER_TO_COL = {
    "Date Applied":   "date_applied",
    "Cover Letter":   "cover_letter",
    "Due Date":       "due_date",
    "Round #":        "round_num",
    "Status":         "status",
    "As of":          "as_of",
    "Notes":          "notes",
    "Application ID": "application_id",
}

# Workbook header -> CSV field, derived from export_workbook.COLUMNS so it
# tracks the workbook schema automatically.
PIPELINE_HEADER_TO_FIELD = {
    header: field for header, field, _ in COLUMNS
    if field is not None and field != "."
}


def _check_schema_assumptions():
    """Fail at import time if export_workbook's COLUMNS drifted away from what
    this mapping assumes — a silent mismatch would import blank user columns
    and quietly lose the edits this whole script exists to preserve."""
    missing = [h for h in USER_HEADER_TO_COL if h not in HEADERS]
    if missing:
        raise RuntimeError(
            f"workbook headers changed — not in export_workbook.COLUMNS: {missing}")
    if set(USER_HEADER_TO_COL.values()) != set(db.USER_FIELDS):
        raise RuntimeError(
            "USER_HEADER_TO_COL and db.USER_FIELDS disagree: "
            f"{set(USER_HEADER_TO_COL.values()) ^ set(db.USER_FIELDS)}")


_check_schema_assumptions()


# ── workbook reading ──────────────────────────────────────────────────────────

def read_workbook(path):
    """Workbook rows as {header: value} dicts, keyed by header NAME so added or
    reordered columns are tolerated (same approach as prune_workbook).

    data_only=False is required: the Website column holds a =HYPERLINK formula,
    and data_only=True would hand back the cached text "Link" instead of the
    URL every row is keyed on.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for cells in ws.iter_rows(min_row=2):
        row = {h: c.value for h, c in zip(header, cells) if h}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    wb.close()
    return rows


def clean(value):
    """Cell -> trimmed string, with the '.' spacer treated as empty."""
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in PLACEHOLDERS else text


def infer_date_processed(value, today=None):
    """Recover a TS_FORMAT timestamp from a workbook 'Date Found' cell.

    Lossy by nature: export_workbook.fmt_date wrote these as "Jun 17" — no year,
    no time. Only reachable for rows that have NO matching CSV row (hand-added
    ones); everything else takes the CSV's full timestamp instead.

    The year is inferred as the most recent one that does not put the date in
    the future, which is right for any tracker holding under a year of history.
    Returns "" if nothing parseable is there, rather than inventing a date.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime(TS_FORMAT)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime(TS_FORMAT)

    text  = str(value).strip()
    today = today or date.today()

    # Full ISO date, if the user retyped one.
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").strftime(TS_FORMAT)
    except ValueError:
        pass

    # export_workbook's "Jun 17" style.
    try:
        parsed = datetime.strptime(text, "%b %d")
    except ValueError:
        return ""
    year = today.year
    if (parsed.month, parsed.day) > (today.month, today.day):
        year -= 1                      # hasn't happened yet this year -> last year
    return datetime(year, parsed.month, parsed.day).strftime(TS_FORMAT)


def workbook_row_to_csv_shape(row):
    """Workbook row -> a dict shaped like a matched_jobs.csv row, so db.from_csv_row
    does the coercion. Only used for rows with no CSV counterpart."""
    out = {}
    for header, field in PIPELINE_HEADER_TO_FIELD.items():
        value = row.get(header)
        if header == "Website":
            value = strip_hyperlink(value)
        out[field] = clean(value)
    out["date_processed"] = infer_date_processed(row.get("Date Found"))
    return out


def user_fields_from_row(row):
    """The hand-typed columns worth carrying over. Empty values are omitted so
    the DB keeps NULLs rather than a table full of empty strings."""
    fields = {}
    for header, col in USER_HEADER_TO_COL.items():
        value = clean(row.get(header))
        if value:
            fields[col] = value
    return fields


def workbook_key(row):
    return row_key(strip_hyperlink(row.get("Website")),
                   row.get("Job Title"), row.get("Company"))


# ── sources ───────────────────────────────────────────────────────────────────

def pull_workbook(dest):
    """Fetch the laptop's workbook. Aborts on failure — falling back to the Pi's
    local copy would silently seed the tracker from a stale snapshot missing
    every edit made since the last successful sync, which is precisely the data
    this import exists to rescue."""
    cfg = load_local_config()
    print(f"Pulling {XLSX_PATH.name} from {cfg['remote_host']} ...")
    result = scp([f"{remote_base(cfg)}{XLSX_PATH.name}", str(dest)])
    if result.returncode != 0 or not dest.exists():
        sys.exit(
            f"\nCould not pull {XLSX_PATH.name} from the laptop.\n"
            f"  * is the laptop on the tailnet? (tailscale status)\n"
            f"  * is the workbook open in Excel with a lock held?\n"
            f"Refusing to bootstrap from the Pi's local copy — it may be stale.\n"
            f"Use --workbook <path> to import a file you have already verified."
        )
    return dest


def load_pruned_keys():
    if not PRUNED_KEYS_PATH.exists():
        return set()
    with open(PRUNED_KEYS_PATH, encoding="utf-8") as f:
        return {ln.strip() for ln in f if ln.strip()}


# ── import ────────────────────────────────────────────────────────────────────

def build_plan(wb_rows, csv_rows, pruned_keys):
    """Decide what becomes live and what becomes a tombstone. Pure — no writes —
    so --dry-run reports exactly what --apply will do."""
    csv_by_key = {}
    for row in csv_rows:
        key = row_key(row.get("url"), row.get("title"), row.get("company"))
        if key:
            csv_by_key[key] = row          # last occurrence wins; CSV is append-only

    live, skipped, no_csv_match = [], 0, 0
    seen = set()
    for row in wb_rows:
        key = workbook_key(row)
        if not key or key in seen:
            skipped += 1                   # blank or duplicate workbook row
            continue
        seen.add(key)

        csv_row = csv_by_key.get(key)
        if csv_row:
            rec = db.from_csv_row(csv_row)         # full fidelity: real timestamp
        else:
            rec = db.from_csv_row(workbook_row_to_csv_shape(row))
            no_csv_match += 1                      # hand-added row, date inferred
        live.append((rec, user_fields_from_row(row)))

    tombs_csv   = [db.from_csv_row(r) for k, r in csv_by_key.items() if k not in seen]
    tombs_prune = sorted(pruned_keys - seen - {r["key"] for r in tombs_csv})

    return {
        "live":         live,
        "tombs_csv":    tombs_csv,
        "tombs_prune":  tombs_prune,
        "skipped":      skipped,
        "no_csv_match": no_csv_match,
    }


def apply_plan(conn, plan):
    stamp = db.now_iso()
    counts = {"live": 0, "tombs_csv": 0, "tombs_prune": 0}

    # One transaction: a failure part-way leaves no half-seeded tracker behind.
    with db.transaction(conn):
        for rec, user_fields in plan["live"]:
            if db.insert_new(conn, rec, user_fields=user_fields):
                counts["live"] += 1

        for rec in plan["tombs_csv"]:
            if db.insert_new(conn, rec, deleted_at=stamp, deleted_reason="import-csv"):
                counts["tombs_csv"] += 1

        for key in plan["tombs_prune"]:
            # Pruned keys are bare strings — no title/company survives in that
            # file. The row exists only to occupy the key so the pipeline's
            # ON CONFLICT DO NOTHING refuses to re-add the listing.
            if db.insert_new(conn, {"key": key},
                             deleted_at=stamp, deleted_reason="import-prune"):
                counts["tombs_prune"] += 1

        db.set_meta(conn, "bootstrapped_at", stamp)
    return counts


def main():
    ap = argparse.ArgumentParser(description="One-shot import of the Excel tracker into the DB.")
    ap.add_argument("--dry-run", action="store_true", help="Report what would be imported, write nothing")
    ap.add_argument("--workbook", type=Path, default=None,
                    help="Import this local .xlsx instead of pulling from the laptop")
    args = ap.parse_args()

    conn = db.connect()
    db.init_db(conn)

    if db.is_bootstrapped(conn):
        when = db.get_meta(conn, "bootstrapped_at", "unknown")
        counts = db.counts(conn)
        sys.exit(
            f"Already bootstrapped ({when}) — {counts['total']} rows present.\n"
            f"Re-importing would overwrite live edits with a stale laptop snapshot, "
            f"so there is no --force.\n"
            f"If this import genuinely needs redoing, delete {db.DB_PATH} and rerun."
        )

    # read_workbook loads everything into memory, so the pulled file can go as
    # soon as it is parsed — no cleanup to thread through the exit paths below.
    tmpdir = None
    try:
        if args.workbook:
            if not args.workbook.exists():
                sys.exit(f"No such workbook: {args.workbook}")
            wb_path = args.workbook
        else:
            tmpdir  = tempfile.TemporaryDirectory()
            wb_path = pull_workbook(Path(tmpdir.name) / XLSX_PATH.name)
        wb_label = "laptop" if tmpdir else str(wb_path)
        wb_rows  = read_workbook(wb_path)
    finally:
        if tmpdir:
            tmpdir.cleanup()

    csv_rows = read_matches(CSV_PATH)
    pruned   = load_pruned_keys()
    plan     = build_plan(wb_rows, csv_rows, pruned)

    print(f"\n  workbook rows      {len(wb_rows):>6}   ({wb_label})")
    print(f"  csv rows           {len(csv_rows):>6}")
    print(f"  pruned keys        {len(pruned):>6}")
    print(f"\n  -> live            {len(plan['live']):>6}")
    print(f"  -> tombstone (csv) {len(plan['tombs_csv']):>6}   past hand-deletions")
    print(f"  -> tombstone (prune){len(plan['tombs_prune']):>5}   from pruned_keys.txt")
    if plan["no_csv_match"]:
        print(f"\n  note: {plan['no_csv_match']} workbook row(s) had no CSV match — "
              f"pipeline fields came from the workbook and Date Found's year was inferred.")
    if plan["skipped"]:
        print(f"  note: {plan['skipped']} workbook row(s) skipped (blank or duplicate key).")

    if args.dry_run:
        print("\nDry run — nothing written. Re-run without --dry-run to import.")
        return

    counts = apply_plan(conn, plan)
    final  = db.counts(conn)
    print(f"\nImported: {counts['live']} live, "
          f"{counts['tombs_csv'] + counts['tombs_prune']} tombstoned.")
    print(f"{db.DB_PATH}: {final['total']} rows ({final['live']} live, "
          f"{final['deleted']} tombstoned)")
    print("\nSanity-check against the old pipeline before trusting it:")
    print("    python3 scripts/jobs_left.py")


if __name__ == "__main__":
    main()
