#!/usr/bin/env python3
"""
Build/maintain matched_jobs.xlsx — a job-application tracker styled like a
hand-kept spreadsheet, fed from the pipeline's matched_jobs.csv.

Append-only by design. New matches are added as new rows; rows already in the
workbook are never rewritten, so the columns you fill in by hand (Status,
Notes, dates, ...) survive every run. Dedup is by Website URL (fallback
title|company), mirroring filter_jobs.job_fingerprint.

Usage:
    python3 export_workbook.py
    python3 export_workbook.py --csv matched.csv --out tracker.xlsx
    python3 export_workbook.py --no-color
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("export_workbook.py needs openpyxl — run: pip install openpyxl")

# row_key moved to matches.py (stdlib only) so db.py and the web app can key
# rows without importing openpyxl. Re-exported here: prune_workbook.py and
# other callers still do `from export_workbook import row_key`.
from matches import month_day, read_matches, row_key  # noqa: F401 — re-export
from paths import (DATA_DIR, EXPORT_MARK_PATH as MARK_PATH,
                   EXPORT_MARK_PENDING as MARK_PENDING)

CSV_PATH  = DATA_DIR / "matched_jobs.csv"
XLSX_PATH = DATA_DIR / "matched_jobs.xlsx"
# Dedup keys of rows trimmed by a manual `prune_workbook.py --apply` run.
# Skipped on future appends so the cumulative matched_jobs.csv never re-adds a
# row that was intentionally pruned. Delete this file if you want previously-
# pruned listings to be reconsidered.
PRUNED_KEYS_PATH = DATA_DIR / "pruned_keys.txt"

# Last run's held count (see the rollback warning in main). Not part of the
# watermark commit handshake — purely a baseline so the warning fires on a
# sudden JUMP in held rows, not on the standing total of all deletions ever.
HELD_COUNT_PATH = DATA_DIR / "held_count.txt"

# Sync watermark: the date_processed of the newest CSV row that has made it
# into a workbook the laptop confirmed receiving. Any OLDER row that is
# missing from the pulled workbook can only be missing because the user
# deleted it by hand — so it is never re-added. Newer rows are new matches
# and are always appended. The export writes the candidate value to
# MARK_PENDING (paths.py); the orchestrator promotes it to MARK_PATH only
# after a successful push, so a failed push never strands unexported rows
# behind the watermark. Timestamps are matches.TS_FORMAT (UTC) — the lexical
# ts <= mark comparison below is only a valid time ordering because that
# format is fixed-width, zero-padded, most-significant-first.

MAX_VALIDATION_ROW = 5000

# ── column layout ─────────────────────────────────────────────────────────────
# (header, csv_field or None for manual columns, width)
COLUMNS = [
    ("Score",          "score",          8),
    ("Job Title",      "title",          50),
    ("Company",        "company",        20),
    ("Location",       "location",       18),
    ("Pay",            "salary",         12),
    ("Website",        "url",            16),
    ("Date Found",     "date_processed", 16),
    ("Date Applied",   None,             16),
    ("Why",            "reason",         45),
    ("Matched Skills", "matched_skills", 30),
    ("Concerns",       "concerns",       30),
    ("Application ID", ".",              16),   # auto-filled with "." to block overflow
    ("Cover Letter",   None,             16),
    ("Due Date",       None,             12),
    ("Round #",        None,             10),
    ("Status",         None,             18),
    ("As of",          None,             12),
    ("Notes",          None,             30),
]
HEADERS     = [c[0] for c in COLUMNS]
WEBSITE_COL = HEADERS.index("Website") + 1
TITLE_COL   = HEADERS.index("Job Title") + 1
COMPANY_COL = HEADERS.index("Company") + 1

# Option lists live in db.py (stdlib) so the web UI's <select>s and this
# workbook's data-validation dropdowns share one definition. Header-keyed here
# for the data-validation loop below; db keys them by jobs-table column name.
from db import USER_FIELD_OPTIONS  # noqa: E402

DROPDOWNS = {
    "Cover Letter": USER_FIELD_OPTIONS["cover_letter"],
    "Status":       USER_FIELD_OPTIONS["status"],
}

HEADER_BG = "FF1F3864"
HEADER_FG = "FFFFFFFF"


# ── helpers ───────────────────────────────────────────────────────────────────

def fmt_date(val):
    """Format date_processed to clean 'Jun 17' style."""
    if not val:
        return ""
    try:
        from datetime import datetime
        return month_day(datetime.strptime(str(val).strip()[:10], "%Y-%m-%d"))
    except Exception:
        return str(val)[:10]


def csv_row_to_cells(row):
    cells = []
    for _, field, _ in COLUMNS:
        if field is None:
            cells.append("")
        elif field == ".":
            cells.append(".")
        elif field == "score":
            try:
                cells.append(int(float(row.get("score", "") or 0)))
            except (TypeError, ValueError):
                cells.append(row.get("score", ""))
        elif field == "date_processed":
            cells.append(fmt_date(row.get("date_processed", "")))
        else:
            cells.append(row.get(field, ""))
    return cells


def style_website_cell(cell):
    """Use HYPERLINK formula — avoids openpyxl's broken inline xmlns:r on hyperlinks."""
    url = (cell.value or "").strip()
    if url:
        safe       = url.replace('"', '%22')
        cell.value = f'=HYPERLINK("{safe}","Link")'
        cell.font  = Font(color="0563C1", underline="single")


def strip_hyperlink(value):
    """Raw URL from a Website cell, which style_website_cell wrote as
    =HYPERLINK("url","Link"). Needed anywhere the workbook is read back —
    export's dedup and bootstrap_from_workbook's import both key on the URL.

    Only correct when the workbook is loaded with data_only=False (the
    default): with data_only=True openpyxl returns the cached display text
    ("Link") and the URL is simply not there to recover.
    """
    if isinstance(value, str) and value.startswith("=HYPERLINK"):
        try:
            return value.split('"')[1]
        except IndexError:
            pass
    return value


def existing_keys(ws):
    keys = set()
    for r in range(2, ws.max_row + 1):
        website = strip_hyperlink(ws.cell(row=r, column=WEBSITE_COL).value)
        title   = ws.cell(row=r, column=TITLE_COL).value
        company = ws.cell(row=r, column=COMPANY_COL).value
        if website or title or company:
            keys.add(row_key(website, title, company))
    return keys


def load_pruned_keys():
    """Dedup keys of rows previously trimmed by prune_workbook()."""
    if not PRUNED_KEYS_PATH.exists():
        return set()
    with open(PRUNED_KEYS_PATH, encoding="utf-8") as f:
        return {ln.strip() for ln in f if ln.strip()}


# (append_pruned_keys lives in prune_workbook.py — the manual tool that owns
#  writing the suppress list; this module only reads it.)


def load_mark():
    """Committed sync watermark, or None if no sync has been confirmed yet."""
    if MARK_PATH.exists():
        return MARK_PATH.read_text(encoding="utf-8").strip()
    return None


def bootstrap_mark(matches, accounted_keys):
    """First run without a watermark: everything already in the workbook (or
    already pruned) is accounted for; the newest such timestamp separates
    hand-deleted history (older, absent) from not-yet-synced matches (newer)."""
    return max((r.get("date_processed", "") or "" for r in matches
                if row_key(r.get("url"), r.get("title"), r.get("company"))
                in accounted_keys), default="")


# ── workbook creation ─────────────────────────────────────────────────────────

def create_workbook(use_color):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"

    # Header row with dark navy styling
    ws.append(HEADERS)
    for i, hdr in enumerate(HEADERS, start=1):
        cell            = ws.cell(row=1, column=i)
        cell.font       = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
        cell.fill       = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment  = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Auto-filter (filter dropdowns without the problematic Table XML)
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # Column widths
    for i, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    _apply_validations(ws)
    if use_color:
        _apply_score_color(ws)

    return wb, ws


def _apply_validations(ws):
    for header, values in DROPDOWNS.items():
        col = get_column_letter(HEADERS.index(header) + 1)
        dv  = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{MAX_VALIDATION_ROW}")


def _apply_score_color(ws):
    col  = get_column_letter(HEADERS.index("Score") + 1)
    rule = ColorScaleRule(
        start_type="num", start_value=0,  start_color="FF6B6B",
        mid_type="num",   mid_value=5,    mid_color="FFD966",
        end_type="num",   end_value=10,   end_color="7AD151",
    )
    ws.conditional_formatting.add(f"{col}2:{col}{MAX_VALIDATION_ROW}", rule)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Build/append the job-tracker .xlsx from matched_jobs.csv.")
    ap.add_argument("--csv",      default=str(CSV_PATH))
    ap.add_argument("--out",      default=str(XLSX_PATH))
    ap.add_argument("--no-color", action="store_true")
    args = ap.parse_args()

    # A pending watermark must only ever come from THIS run. Clear any stale
    # one (e.g. from a cycle whose push failed) BEFORE the early returns below,
    # or the orchestrator could promote it after an unrelated successful push —
    # committing a watermark the pushed workbook never reflected.
    MARK_PENDING.unlink(missing_ok=True)

    matches = read_matches(Path(args.csv))
    if not matches:
        print(f"No rows in {args.csv} — nothing to export.")
        return

    out = Path(args.out)
    fresh_workbook = not out.exists()
    if fresh_workbook:
        wb, ws = create_workbook(use_color=not args.no_color)
        seen   = set()
    else:
        wb   = load_workbook(out)
        ws   = wb["Matches"] if "Matches" in wb.sheetnames else wb.active
        seen = existing_keys(ws)

    # Don't re-add rows a manual prune has trimmed (the CSV is cumulative).
    seen |= load_pruned_keys()

    # Sync watermark: rows older than the last confirmed sync that are missing
    # from the workbook were deleted by hand — leave them deleted. A fresh
    # workbook is a full rebuild, so the watermark doesn't apply (the CSV is
    # the only record we have); pruned keys still suppress.
    if fresh_workbook:
        mark = ""
    else:
        mark = load_mark()
        if mark is None:
            mark = bootstrap_mark(matches, seen)
            if mark:
                print(f"No sync watermark yet — CSV rows up to {mark} that are "
                      f"missing from the workbook stay deleted.")

    added, held = 0, 0
    new_mark = mark
    for row in matches:
        key = row_key(row.get("url"), row.get("title"), row.get("company"))
        ts  = (row.get("date_processed") or "").strip()
        if key in seen:
            new_mark = max(new_mark, ts)   # accounted for (present or pruned)
            continue
        if ts and mark and ts <= mark:
            held += 1                      # pre-watermark and absent → hand-deleted
            continue
        seen.add(key)
        ws.append(csv_row_to_cells(row))
        style_website_cell(ws.cell(row=ws.max_row, column=WEBSITE_COL))
        added += 1
        new_mark = max(new_mark, ts)

    # Update auto_filter range to cover all data rows
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    wb.save(out)
    # Candidate watermark — committed to MARK_PATH by the orchestrator only
    # after the laptop confirms the push, so a failed push re-appends next time.
    MARK_PENDING.write_text(new_mark, encoding="utf-8")
    print(f"{added} new row(s) added; {held} hand-deleted row(s) left deleted; "
          f"workbook now has {ws.max_row - 1} matches -> {out}")
    # held is a STANDING total (hand-deleted rows stay in the cumulative CSV
    # forever), so warn only when it JUMPS — dozens of rows newly held in one
    # run is the signature of a laptop workbook restored from an older backup,
    # not of gradual hand-pruning. Deleting the mark file re-adds everything
    # still in the CSV.
    prev_held = 0
    try:
        prev_held = int(HELD_COUNT_PATH.read_text(encoding="utf-8").strip() or 0)
    except (OSError, ValueError):
        pass
    HELD_COUNT_PATH.write_text(str(held), encoding="utf-8")
    if held - prev_held >= 20:
        print(f"WARNING: held rows jumped {prev_held} -> {held} in one run. If "
              f"the laptop workbook was restored from a backup, delete "
              f"{MARK_PATH} and re-run to recover the missing rows. (If you "
              f"really did just hand-delete {held - prev_held} rows, ignore this.)")


if __name__ == "__main__":
    main()
