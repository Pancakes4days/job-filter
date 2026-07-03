#!/usr/bin/env python3
"""
prune_workbook.py — trim the job-tracker workbook down to the best 1–2 jobs
per company, protecting rows you've started applying to.

MANUAL TOOL — run it by hand when the workbook gets noisy. The automated
pipeline (export_workbook.py) is purely append-only and never deletes rows.

    python3 scripts/prune_workbook.py                  # dry-run report, no writes
    python3 scripts/prune_workbook.py --apply          # pull → prune → push
    python3 scripts/prune_workbook.py --apply --local  # prune the local file only

--apply pulls the laptop's workbook FIRST (that copy holds your hand edits),
prunes, pushes the result back, and only after a confirmed push records the
deleted rows' dedup keys in data/pruned_keys.txt (so the cumulative
matched_jobs.csv never re-adds them). If the laptop can't be reached it aborts
rather than prune a stale copy; if the push fails, no keys are recorded, so
state stays consistent. The dry run pulls the laptop's copy to a temp file for
an accurate preview when possible. --local prunes of the live workbook are
transient (the next sync restores the laptop's copy) and record no keys.

Design notes:
  * prune_workbook() mutates the worksheet in place and returns the dedup keys
    of the rows it deleted; file/network I/O lives only in the CLI below.
  * Columns are matched by HEADER NAME, so column reordering/additions are tolerated.
  * Rows with a real user value in any PROTECT column are never deleted. The
    pipeline auto-fills "Application ID" with "." as a spacer, so "." counts as empty.

Candidate profile baked into the rules: May 2027 grad targeting Summer 2027
internships and new-grad roles. Strength order: security > Applied AI/ML >
DevOps/SRE/infra > backend/full-stack. Cuts senior/2+yr/PhD/quant/non-engineering
and unacceptable-location rows. Edit the keyword lists below to retune.
"""
import argparse
import os
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

# export_workbook exits with a pip hint if openpyxl is missing, so import it
# before openpyxl. It owns the workbook schema (row_key) and the pipeline's
# pruned-keys suppress list location.
from export_workbook import PRUNED_KEYS_PATH, XLSX_PATH, row_key
from openpyxl import load_workbook
from paths import DATA_DIR
from remote import LOCAL_JSON, load_local_config, remote_base, scp

KEEP_PER_COMPANY = 2

# User-edited columns. A real value in any of these protects the row from deletion.
PROTECT_COLS = ["Date Applied", "Application ID", "Cover Letter",
                "Due Date", "Round #", "Status", "As of", "Notes"]
# "Application ID" is auto-filled with "." by export_workbook.py as a spacer; "" and
# "." both count as empty so they never falsely protect a row.
PLACEHOLDERS = {"", "."}

# ── location ──────────────────────────────────────────────────────────────────
ACCEPTABLE_LOC = ["new york", "nyc", "long island", "manhattan", "brooklyn",
                  "florida", "virginia", "boston", "massachusetts", "colorado",
                  "denver", "switzerland", "zurich", "riva", "spain",
                  "san francisco", "bay area", "palo alto", "remote",
                  "united states"]
NYC = ["new york", "nyc", "long island", "manhattan", "brooklyn"]
MID = ["florida", "virginia", "boston", "massachusetts", "colorado", "denver",
       "switzerland", "zurich", "riva", "spain"]

# ── exclusion patterns ────────────────────────────────────────────────────────
SENIOR = re.compile(r"\b(senior|sr\.?|staff|principal|\blead\b|manager|director|"
                    r"\bhead\b|\bvp\b|vice president|chief|architect|fellow)\b", re.I)
LEVEL3 = re.compile(r"engineer\s*(3|iii)\b", re.I)
YEARS  = re.compile(r"(\d+)\s*\+?\s*(?:[-–]\s*\d+\s*)?\s*(?:years|yrs|yoe)\b", re.I)
PHD_TITLE = re.compile(r"\bph\.?\s*d\b", re.I)
# New-grad/intern signal must be in the TITLE to override a senior title. The LLM's
# Why/Concerns text says "new grad" in NEGATIVE contexts ("outside the new grad
# timeframe"), so a blob match is not a reliable rescue.
NEWGRAD_TITLE = re.compile(r"new\s*grad|entry[- ]level|\bintern(ship)?\b|"
                           r"university (grad|program)|graduate (program|engineer|rotational)|"
                           r"\bgrad program\b|2027 grad|co[- ]?op\b", re.I)
WRONGCYCLE = re.compile(r"grad\w*\s+before\s+may\s+2027|before may 2027|"
                        r"fall\s*2026\s*start|summer\s*2026|experienced professionals", re.I)
DEGREE = re.compile(r"\b(ms|m\.s\.|master'?s|phd|ph\.d\.)\b[^.]*\brequired", re.I)
# HARD non-engineering: never an SWE role even if the title contains "engineer".
HARD_NONENG = re.compile(
    r"account executive|account manager|\bsales\b|sales engineer|solutions? engineer|"
    r"business development|partner development|\brepresentative\b|\brecruiter\b|\bsourcer\b|\btalent\b|"
    r"\battorney\b|\bcounsel\b|\blegal\b|\bmarketing\b|product designer|gtm engineer|"
    r"developer advocate|\badvocate\b|community growth|revenue enablement|\benablement\b|"
    r"\bconsultant\b|customer success|customer experience|people ops|ai trainer", re.I)
# SOFT non-engineering: non-eng UNLESS the title is a genuine engineering IC role.
SOFT_NONENG = re.compile(
    r"\banalyst\b|\bscientist\b|\bresearcher\b|\bassociate\b|\boperations\b|"
    r"\bcoordinator\b|\bspecialist\b|\bstrategist\b|\btrader\b|\btrading\b|\btrainer\b|"
    r"equity research|research analyst|data scientist|applied scientist|research scientist|"
    r"\bdesigner\b|\bfinance\b|fp&a|\binvestment\b|treasury|finops|fin ops|"
    r"business analyst|business systems|business automation|ops analyst|"
    r"\bbilling\b|accounts receivable|\bproducer\b|\beducator\b|\bcompliance\b|"
    r"product manage|program manage|customer data|\bsalesforce\b|"
    r"support engineer|technical support|solutions? architect", re.I)
SWE_OK = re.compile(r"software engineer|design engineer|full[- ]?stack|backend|"
                    r"front[- ]?end|web application", re.I)
QUANT = re.compile(r"quantitative (trader|researcher|analyst)|quant researcher", re.I)
EXPLICIT_2027 = re.compile(r"2027|summer 2027|2027 grad", re.I)

# ── fit tiers (lower = better) ───────────────────────────────────────────────
T_AI  = re.compile(r"applied ai|ai engineer|machine learning|\bml\b|ml ops|mlops|"
                   r"\bnlp\b|\bllm\b|gen ?ai|ai platform", re.I)
T_SWE = re.compile(r"software engineer|full[- ]?stack|backend|front[- ]?end|"
                   r"web application|developer", re.I)
T_INF = re.compile(r"devops|platform engineer|site reliability|\bsre\b|infrastructure|"
                   r"reliability engineer|linux engineer|environment platform|"
                   r"apollo|cloud operations", re.I)
T_SEC = re.compile(r"security|appsec|infosec|vulnerability", re.I)


def _title_newgrad(title):
    return bool(NEWGRAD_TITLE.search(title))


def _eng_ok(title):
    # genuine engineering IC role -> rescue from a SOFT non-engineering match
    return bool(SWE_OK.search(title) or T_AI.search(title) or T_SWE.search(title)
                or T_INF.search(title) or T_SEC.search(title))


def _is_senior(title):
    tl = title.lower()
    # "Member of Technical Staff" is an IC title at AI labs, not a senior level.
    chk = tl.replace("technical staff", "") if "technical staff" in tl else title
    return bool(SENIOR.search(chk)) or bool(LEVEL3.search(title))


def _hard_excluded(title, location, why, concerns):
    blob = " ".join([title, why, concerns])
    ng = _title_newgrad(title)
    loc = location.lower()
    if WRONGCYCLE.search(blob): return "wrong cycle"
    if PHD_TITLE.search(title): return "PhD role"
    if _is_senior(title) and not ng: return "too senior"
    m = YEARS.search(title)
    if m and int(m.group(1)) >= 2 and not ng: return "needs 2+ yrs"
    if DEGREE.search(blob): return "advanced degree required"
    if HARD_NONENG.search(title): return "non-engineering"
    if SOFT_NONENG.search(title) and not _eng_ok(title): return "non-engineering"
    if QUANT.search(title) and not re.search(r"developer|intern|new grad", title, re.I):
        return "quant specialist"
    if not any(k in loc for k in ACCEPTABLE_LOC): return "location"
    return None


def _fit_tier(title):
    if T_SEC.search(title):  return 0
    if T_AI.search(title): return 1
    if T_INF.search(title): return 2
    if T_SWE.search(title): return 3
    return 4


def _loc_rank(location):
    loc = location.lower()
    if any(k in loc for k in NYC): return 0
    if any(k in loc for k in MID): return 1
    return 2


def _score_val(s):
    try: return float(s)
    except (TypeError, ValueError): return 0.0


def prune_workbook(ws, row_key):
    """Trim `ws` (a worksheet with a header row) to the best 1–2 jobs per company.

    `row_key(website, title, company)` is the caller's dedup-key function; it is used
    to build the list of deleted-row keys returned for the suppress-list.

    Protected rows (any real value in PROTECT_COLS) are never deleted. Returns
    (deleted_keys, kept_count) where deleted_keys is a list of dedup keys.
    """
    header = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(header)}     # 0-based into the cell tuple

    def g(cells, name):
        i = col.get(name)
        v = cells[i].value if i is not None else None
        return "" if v is None else str(v)

    def key_of(cells):
        website = g(cells, "Website")
        if website.startswith("=HYPERLINK"):            # unwrap =HYPERLINK("url","Link")
            try: website = website.split('"')[1]
            except IndexError: pass
        return row_key(website, g(cells, "Job Title"), g(cells, "Company"))

    def is_protected(cells):
        return any(g(cells, c).strip().lower() not in PLACEHOLDERS for c in PROTECT_COLS)

    def sort_key(cells):
        title = g(cells, "Job Title")
        blob = " ".join([title, g(cells, "Why"), g(cells, "Concerns")])
        cyc = 0 if EXPLICIT_2027.search(blob) else 1
        return (_fit_tier(title), _loc_rank(g(cells, "Location")), cyc,
                -_score_val(g(cells, "Score")))

    # gather data rows with their 1-based sheet index
    rows = [(idx, cells) for idx, cells in enumerate(ws.iter_rows(min_row=2), start=2)]
    by_company = defaultdict(list)
    for idx, cells in rows:
        by_company[g(cells, "Company")].append((idx, cells))

    delete_idx = []
    deleted_keys = []
    kept = 0
    for _company, items in by_company.items():
        protected = [(i, c) for i, c in items if is_protected(c)]
        cand      = [(i, c) for i, c in items if not is_protected(c)]
        survivors = [(i, c) for i, c in cand
                     if not _hard_excluded(g(c, "Job Title"), g(c, "Location"),
                                           g(c, "Why"), g(c, "Concerns"))]
        survivors.sort(key=lambda ic: sort_key(ic[1]))
        keep = survivors[:KEEP_PER_COMPANY]
        keep_idx = {i for i, _ in protected} | {i for i, _ in keep}
        kept += len(keep_idx)
        for i, c in items:
            if i not in keep_idx:
                delete_idx.append(i)
                deleted_keys.append(key_of(c))

    for i in sorted(delete_idx, reverse=True):           # bottom-up so indices hold
        ws.delete_rows(i, 1)

    return deleted_keys, kept


# ── standalone CLI ────────────────────────────────────────────────────────────

def append_pruned_keys(keys):
    """Record deleted rows' dedup keys in the pipeline's suppress list so the
    cumulative matched_jobs.csv never re-adds them. Only call this once the
    pruned workbook is authoritative (pushed to the laptop) — recording keys
    for rows that still exist somewhere just creates inconsistent state."""
    if not keys:
        return
    with open(PRUNED_KEYS_PATH, "a", encoding="utf-8") as f:
        for k in keys:
            f.write(k + "\n")


def main():
    ap = argparse.ArgumentParser(
        description="Trim the tracker to the best 1–2 roles per company "
                    "(rows with hand-typed values are never deleted).")
    ap.add_argument("workbook", nargs="?", default=str(XLSX_PATH),
                    help=f"Workbook path (default: {XLSX_PATH}); "
                         f"a custom path implies --local")
    ap.add_argument("--apply", action="store_true",
                    help="Write the changes (default is a dry-run report)")
    ap.add_argument("--local", action="store_true",
                    help="Skip the laptop pull/push; act on the local file only")
    args = ap.parse_args()

    path = Path(args.workbook)
    using_default = args.workbook == str(XLSX_PATH)
    local_only = args.local or not using_default

    if args.apply and local_only and using_default:
        print("WARNING: a --local prune of the live workbook is TRANSIENT — the next\n"
              "         sync pulls the laptop's copy back over it, and no suppress\n"
              "         keys are recorded. Drop --local to prune for real.")

    remote_dir = None
    tmp_pull   = None
    if not local_only:
        if not LOCAL_JSON.exists():
            if args.apply:
                sys.exit(f"Missing {LOCAL_JSON} — use --local to prune the "
                         f"local file only.")
            print(f"No {LOCAL_JSON.name} — previewing the local copy.")
            local_only = True
        else:
            remote_dir = remote_base(load_local_config())
            if args.apply:
                # The laptop's copy holds your hand edits — prune THAT, never
                # a stale one.
                print(f"Pulling {path.name} from the laptop…")
                if scp([remote_dir + path.name, str(path)]).returncode != 0:
                    sys.exit("Couldn't pull the workbook (laptop offline, or file "
                             "open in Excel?) — aborting so nothing is lost. Use "
                             "--local to prune the Pi's copy anyway.")
            else:
                # Dry run: preview the copy --apply would actually prune, in a
                # temp file so the dry run has zero side effects.
                fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(DATA_DIR))
                os.close(fd)
                tmp_pull = Path(tmp)
                print("Pulling the laptop's workbook for an accurate preview…")
                if scp([remote_dir + path.name, str(tmp_pull)]).returncode == 0:
                    path = tmp_pull
                else:
                    tmp_pull.unlink(missing_ok=True)
                    tmp_pull = None
                    print("  couldn't pull — previewing the LOCAL copy instead "
                          "(may miss recent hand edits).")

    if not path.exists():
        sys.exit(f"No workbook at {path}")

    wb = load_workbook(path)
    ws = wb["Matches"] if "Matches" in wb.sheetnames else wb.active
    before = ws.max_row - 1
    deleted, _kept = prune_workbook(ws, row_key)
    print(f"  data rows: {before} -> {before - len(deleted)} "
          f"(delete {len(deleted)}; hand-edited rows are protected)")

    if not args.apply:
        if tmp_pull:
            tmp_pull.unlink(missing_ok=True)
        print("Dry run — nothing written. Re-run with --apply to prune for real.")
        return

    wb.save(path)
    print(f"Pruned {len(deleted)} row(s) from {path.name}.")

    if local_only:
        # No suppress keys on purpose: these rows still exist in the laptop
        # workbook / cumulative CSV, so keys would suppress rows that were
        # never authoritatively removed.
        return

    print(f"Pushing {path.name} back to the laptop…")
    if scp([str(path), remote_dir]).returncode != 0:
        print("PUSH FAILED — the laptop still has the unpruned copy and no keys\n"
              "were recorded, so nothing is inconsistent. Re-run --apply once the\n"
              "laptop is reachable.")
        sys.exit(1)
    append_pruned_keys(deleted)
    print(f"Pushed. {len(deleted)} pruned key(s) recorded in "
          f"{PRUNED_KEYS_PATH.name} so the pipeline won't re-add them.")


if __name__ == "__main__":
    main()
